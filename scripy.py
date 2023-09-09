import openpyxl
f= open("output.txt", 'w')
workbook = openpyxl.load_workbook('rules.xlsx')  # Replace 'example.xlsx' with your Excel file's name
sheet = workbook.active
max_row=sheet.max_row
x = "NULL"
policycount = 0
for i in range (2, max_row+1):
    devicename = (sheet.cell(i, 1).value)
    if x == devicename:
        pass
    else:
        f.write("======"+devicename+"======"+"\n")
        x = devicename
    policyname = (sheet.cell(i, 2).value)
    initialsource = (sheet.cell(i, 8).value)
    srcoperation = initialsource.split(",")
    if len(srcoperation) == 1:
        if "/" in initialsource:
            src = initialsource
        else:
            src = "host " + initialsource
    else:
        src = "object-group " + srcoperation[0]

    initialdestination = (sheet.cell(i, 10).value)
    dstoperation = initialdestination.split(",")
    if len(dstoperation) == 1:
        if "/" in initialdestination:
            dst = initialdestination
        else:
            dst = "host " + initialdestination
    else:
        dst = "object-group " + dstoperation[0]
    
    initialservice = (sheet.cell(i, 11).value)
    srvoperation = initialservice.split(",")
    if len(srvoperation) == 1:
        if "tcp" in initialservice:
            protocol = "tcp"
            service = "eq " + initialservice.replace("tcp-","")
            policycount = 1
        elif "udp" in initialservice:
            protocol = "udp"
            service = "eq " + initialservice.replace("udp-","")
            policycount = 1
    else:
        if "tcp" in initialservice and "udp" in initialservice:
            service = "object-group "+ srvoperation[0]
            policycount = 2
        elif "tcp" in initialservice:
            protocol = "tcp"
            service = "object-group "+ srvoperation[0]
            policycount = 1
        elif "udp" in initialservice:
            protocol = "udp"
            service = "object-group "+ srvoperation[0]
            policycount = 1
    ruleaction = (sheet.cell(i, 14).value)
    ruleaction = ruleaction.replace("DROP", "deny")
    ruleaction = ruleaction.replace("ACCEPT", "permit")
    if policycount == 1:
        f.write(f"access-list {policyname} extended {ruleaction} {protocol} {src} {dst} {service} log\n")
    elif policycount == 2:
        f.write(f"access-list {policyname} extended {ruleaction} tcp {src} {dst} {service} log\n")
        f.write(f"access-list {policyname} extended {ruleaction} udp {src} {dst} {service} log\n")
    i = i+1
workbook.close()
